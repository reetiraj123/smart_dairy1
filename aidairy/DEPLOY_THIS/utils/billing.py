"""
Billing utility module for SmartDairy
Handles monthly billing calculations and exports (PDF, Excel, CSV)
"""

import pandas as pd
from datetime import datetime
from typing import List, Dict
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import os
from utils.db import get_monthly_entries

def calculate_monthly_billing(year: int, month: int) -> Dict:
    """
    Calculate monthly billing for all customers
    Returns a dictionary with billing summary
    """
    entries = get_monthly_entries(year, month)
    
    # Group by customer
    customer_totals = {}
    for entry in entries:
        customer_id = entry['customer_id']
        customer_name = entry['customer_name']
        price_per_ltr = entry['price_per_ltr']
        quantity = entry['quantity']
        mobile_number = entry.get('mobile_number', '')
        
        if customer_id not in customer_totals:
            customer_totals[customer_id] = {
                'id': customer_id,
                'name': customer_name,
                'price_per_ltr': price_per_ltr,
                'mobile_number': mobile_number,
                'total_litres': 0.0,
                'total_amount': 0.0
            }
        
        customer_totals[customer_id]['total_litres'] += quantity
        customer_totals[customer_id]['total_amount'] = (
            customer_totals[customer_id]['total_litres'] * price_per_ltr
        )
    
    # Calculate grand total
    grand_total = sum(cust['total_amount'] for cust in customer_totals.values())
    
    return {
        'year': year,
        'month': month,
        'customers': list(customer_totals.values()),
        'grand_total': grand_total,
        'total_customers': len(customer_totals)
    }

def generate_pdf_invoice(billing_data: Dict, output_path: str = "invoice.pdf"):
    """Generate PDF invoice using ReportLab"""
    doc = SimpleDocTemplate(output_path, pagesize=A4)
    story = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#2E86AB'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#1B4332'),
        spaceAfter=12
    )
    
    # Title
    story.append(Paragraph("SmartDairy - Monthly Invoice", title_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Invoice details
    month_name = datetime(billing_data['year'], billing_data['month'], 1).strftime('%B %Y')
    story.append(Paragraph(f"<b>Invoice Period:</b> {month_name}", styles['Normal']))
    story.append(Paragraph(f"<b>Generated On:</b> {datetime.now().strftime('%d %B %Y, %I:%M %p')}", styles['Normal']))
    story.append(Spacer(1, 0.3*inch))
    
    # Customer billing table
    story.append(Paragraph("Billing Summary", heading_style))
    
    # Table data
    table_data = [['Customer Name', 'Total Litres', 'Rate/Litre (₹)', 'Total Amount (₹)']]
    
    for customer in billing_data['customers']:
        table_data.append([
            customer['name'],
            f"{customer['total_litres']:.2f}",
            f"{customer['price_per_ltr']:.2f}",
            f"{customer['total_amount']:.2f}"
        ])
    
    # Grand total row
    table_data.append([
        '<b>TOTAL</b>',
        '',
        '',
        f"<b>₹{billing_data['grand_total']:.2f}</b>"
    ])
    
    # Create table
    table = Table(table_data, colWidths=[3*inch, 1.5*inch, 1.5*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E86AB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -2), colors.black),
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F77F00')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 12),
    ]))
    
    story.append(table)
    story.append(Spacer(1, 0.3*inch))
    
    # Footer
    story.append(Paragraph(
        "<i>This is a computer-generated invoice. Thank you for using SmartDairy!</i>",
        styles['Normal']
    ))
    
    doc.build(story)
    return output_path

def generate_excel_invoice(billing_data: Dict, output_path: str = "invoice.xlsx"):
    """Generate Excel invoice"""
    # Prepare data
    data = []
    for customer in billing_data['customers']:
        data.append({
            'Customer Name': customer['name'],
            'Total Litres': round(customer['total_litres'], 2),
            'Rate per Litre (₹)': round(customer['price_per_ltr'], 2),
            'Total Amount (₹)': round(customer['total_amount'], 2)
        })
    
    # Add grand total row
    data.append({
        'Customer Name': 'GRAND TOTAL',
        'Total Litres': '',
        'Rate per Litre (₹)': '',
        'Total Amount (₹)': round(billing_data['grand_total'], 2)
    })
    
    df = pd.DataFrame(data)
    
    # Write to Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Monthly Invoice', index=False)
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Monthly Invoice']
        
        # Style header row
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Style total row
        total_row = len(data)
        for cell in worksheet[total_row]:
            cell.font = Font(bold=True)
            if cell.column == 4:  # Total Amount column
                cell.fill = PatternFill(start_color="F77F00", end_color="F77F00", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    return output_path

def generate_csv_invoice(billing_data: Dict, output_path: str = "invoice.csv"):
    """Generate CSV invoice"""
    data = []
    for customer in billing_data['customers']:
        data.append({
            'Customer Name': customer['name'],
            'Total Litres': round(customer['total_litres'], 2),
            'Rate per Litre (₹)': round(customer['price_per_ltr'], 2),
            'Total Amount (₹)': round(customer['total_amount'], 2)
        })
    
    # Add grand total
    data.append({
        'Customer Name': 'GRAND TOTAL',
        'Total Litres': '',
        'Rate per Litre (₹)': '',
        'Total Amount (₹)': round(billing_data['grand_total'], 2)
    })
    
    df = pd.DataFrame(data)
    df.to_csv(output_path, index=False)
    return output_path

def format_bill_message(customer: Dict, billing_data: Dict) -> str:
    """Format bill message for WhatsApp"""
    month_name = datetime(billing_data['year'], billing_data['month'], 1).strftime('%B %Y')
    
    message = f"""🐄 *SmartDairy - Monthly Invoice*

*Customer:* {customer['name']}
*Period:* {month_name}

*Billing Details:*
━━━━━━━━━━━━━━━━━━━━
📊 Total Litres: {customer['total_litres']:.2f} L
💰 Rate per Litre: ₹{customer['price_per_ltr']:.2f}
💵 *Total Amount: ₹{customer['total_amount']:.2f}*
━━━━━━━━━━━━━━━━━━━━

Thank you for your business!
_This is an automated message from SmartDairy System_"""
    
    return message

def send_whatsapp_bill(customer: Dict, billing_data: Dict):
    """
    Send bill via WhatsApp using pywhatkit
    Returns: (success: bool, message: str)
    """
    try:
        import pywhatkit as pwk
        from datetime import datetime, timedelta
        
        mobile_number = customer.get('mobile_number', '').strip()
        
        if not mobile_number:
            return False, "Mobile number not found for this customer"
        
        # Remove any non-digit characters except +
        mobile_clean = ''.join(c for c in mobile_number if c.isdigit() or c == '+')
        
        # If no country code, assume Indian number (add +91)
        if not mobile_clean.startswith('+'):
            if len(mobile_clean) == 10:
                mobile_clean = '+91' + mobile_clean
            elif len(mobile_clean) > 10:
                mobile_clean = '+' + mobile_clean
        
        # Format message
        message = format_bill_message(customer, billing_data)
        
        # Get current time + 1 minute (pywhatkit needs time in future)
        now = datetime.now()
        send_time = now + timedelta(minutes=1)
        hour = send_time.hour
        minute = send_time.minute
        
        # Send WhatsApp message
        pwk.sendwhatmsg(mobile_clean, message, hour, minute, wait_time=15, tab_close=True)
        
        return True, f"Bill sent successfully to {customer['name']} at {mobile_clean}"
        
    except ImportError:
        return False, "pywhatkit library not installed. Please install it using: pip install pywhatkit"
    except Exception as e:
        return False, f"Error sending WhatsApp message: {str(e)}"

def get_whatsapp_link(customer: Dict, billing_data: Dict) -> str:
    """Generate WhatsApp web link for manual sending"""
    mobile_number = customer.get('mobile_number', '').strip()
    
    if not mobile_number:
        return ""
    
    # Remove any non-digit characters except +
    mobile_clean = ''.join(c for c in mobile_number if c.isdigit() or c == '+')
    
    # If no country code, assume Indian number (add 91)
    if not mobile_clean.startswith('+'):
        if len(mobile_clean) == 10:
            mobile_clean = '91' + mobile_clean
        elif len(mobile_clean) > 10:
            mobile_clean = mobile_clean
    
    # Format message
    message = format_bill_message(customer, billing_data)
    
    # URL encode message
    from urllib.parse import quote
    encoded_message = quote(message)
    
    # Generate WhatsApp link
    whatsapp_link = f"https://wa.me/{mobile_clean}?text={encoded_message}"
    return whatsapp_link

